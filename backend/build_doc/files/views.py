import os
import io
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse
from django.utils import timezone
from .models import UserFile, FavoriteFile, FileComment, Project, ProjectFile
from .serializers import UserFileSerializer, FileCommentSerializer, ProjectSerializer


class AllFilesListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Можно добавить фильтрацию/поиск/сортировку как в UserFileListView
        files = UserFile.objects.all().order_by('-uploaded_at')

        # Фильтрация по этапу
        stage = request.query_params.get('stage', '')

        if stage:
            files = files.filter(stage=stage)

        # Поиск по названию
        search = request.query_params.get('search', '')

        if search:
            files = files.filter(title__icontains=search)

        # Фильтр по дате
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)

        # Сортировка
        sort = request.query_params.get('sort', '-uploaded_at')
        files = files.order_by(sort)

        serializer = UserFileSerializer(files, many=True, context={'request': request})
        return Response(serializer.data)


class UserFileUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, format=None):
        files = request.FILES.getlist('files')
        titles = request.data.getlist('titles')
        stages = request.data.getlist('stages')
        uploaded_files = []

        if not files:
            return Response({'error': 'Нет файлов для загрузки.'}, status=400)

        for idx, file in enumerate(files):
            title = titles[idx] if idx < len(titles) else file.name
            stage = stages[idx] if idx < len(stages) else 'construction'
            serializer = UserFileSerializer(data={'title': title, 'file': file, 'stage': stage}, context={'request': request})
            
            if serializer.is_valid():
                serializer.save(owner=request.user)
                uploaded_files.append(serializer.data)
            else:
                return Response(serializer.errors, status=400)

        return Response(uploaded_files, status=201)


class UserFileRenameView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            file = UserFile.objects.get(pk=pk)
        except UserFile.DoesNotExist:
            return Response({'error': 'Файл не найден'}, status=404)
        
        new_title = request.data.get('title')
        
        if not new_title or new_title == file.title:
            return Response({'error': 'Новое название не указано или совпадает с текущим'}, status=400)
        
        file.title = new_title
        file.version += 1
        file.last_edited_by = request.user
        file.last_edited_at = timezone.now()
        file.save()
        serializer = UserFileSerializer(file, context={'request': request})
        
        return Response(serializer.data)


class UserFileListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Фильтрация по пользователю
        files = UserFile.objects.filter(owner=request.user)
        
        # Поиск по названию
        search = request.query_params.get('search', '')

        if search:
            files = files.filter(Q(title__icontains=search))
        
        # Фильтр по этапу
        stage = request.query_params.get('stage', '')
        
        if stage:
            files = files.filter(stage=stage)
        

        # Фильтр по дате
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)
        
        # Сортировка
        sort = request.query_params.get('sort', '-uploaded_at')
        files = files.order_by(sort)
        
        serializer = UserFileSerializer(files, many=True, context={'request': request})
        return Response(serializer.data)


class UserFileDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            file = UserFile.objects.get(pk=pk, owner=request.user)
            file.file.delete(save=False)  # Удалить сам файл с диска
            file.delete()

            return Response({'message': 'Файл удалён'}, status=status.HTTP_204_NO_CONTENT)
        except UserFile.DoesNotExist:
            return Response({'error': 'Файл не найден'}, status=status.HTTP_404_NOT_FOUND)


class UserFileDeleteAllView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        files = UserFile.objects.filter(owner=request.user)

        for file in files:
            file.file.delete(save=False)

        files.delete()

        return Response({'message': 'Все файлы удалены'}, status=status.HTTP_204_NO_CONTENT)


class UserFileDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            file_obj = UserFile.objects.get(pk=pk)
            file_path = file_obj.file.path
            
            if not os.path.exists(file_path):
                raise Http404("Файл не найден")
            
            response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(file_path))
            
            return response
        except UserFile.DoesNotExist:
            raise Http404("Файл не найден")


class UserFileDownloadAllView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Если нужно скачивать все файлы (для страницы "Все файлы")
        files = UserFile.objects.all()
        
        if not files.exists():
            raise Http404("Нет файлов для скачивания")
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_obj in files:
                file_path = file_obj.file.path
                if os.path.exists(file_path):
                    # Имя файла в архиве: title + расширение
                    ext = os.path.splitext(file_path)[1]
                    arcname = f"{file_obj.title}{ext}"
                    zip_file.write(file_path, arcname=arcname)
        
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="all_files.zip"'
        
        return response


class UserFileDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            file = UserFile.objects.get(pk=pk)
        except UserFile.DoesNotExist:
            return Response({'error': 'Файл не найден'}, status=404)
        
        serializer = UserFileSerializer(file, context={'request': request})
        
        return Response(serializer.data)


class FileCommentListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, file_id):
        comments = FileComment.objects.filter(file_id=file_id).order_by('created_at')
        serializer = FileCommentSerializer(comments, many=True)
        
        return Response(serializer.data)

    def post(self, request, file_id):
        serializer = FileCommentSerializer(data=request.data)
        
        if serializer.is_valid():
            comment = FileComment.objects.create(
                file_id=file_id,
                author=request.user,
                text=serializer.validated_data['text']
            )
            
            return Response(FileCommentSerializer(comment).data, status=201)
        
        return Response(serializer.errors, status=400)


# Добавить/убрать файл в избранное
class ToggleFavoriteFileView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            file = UserFile.objects.get(pk=pk)
        except UserFile.DoesNotExist:
            return Response({'error': 'Файл не найден'}, status=404)
        
        favorite, created = FavoriteFile.objects.get_or_create(user=request.user, file=file)
        
        if not created:
            favorite.delete()
            return Response({'message': 'Файл удалён из избранного'}, status=200)
        
        return Response({'message': 'Файл добавлен в избранное'}, status=201)


# Получить список избранных файлов пользователя
class MyFavoriteFilesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Получаем id избранных файлов пользователя
        favorite_ids = FavoriteFile.objects.filter(user=request.user).values_list('file_id', flat=True)
        files = UserFile.objects.filter(id__in=favorite_ids)

        # Фильтрация по этапу
        stage = request.query_params.get('stage', '')
        
        if stage:
            files = files.filter(stage=stage)

        # Поиск по названию
        search = request.query_params.get('search', '')
        
        if search:
            files = files.filter(title__icontains=search)

        # Фильтр по дате
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)

        # Сортировка
        sort = request.query_params.get('sort', '-uploaded_at')
        files = files.order_by(sort)

        serializer = UserFileSerializer(files, many=True, context={'request': request})
        
        return Response(serializer.data)


# Скачать все избранные
class DownloadFavoriteFilesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        favorites = FavoriteFile.objects.filter(user=request.user)
        files = [fav.file for fav in favorites]
        
        if not files:
            raise Http404("Нет избранных файлов для скачивания")

        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_obj in files:
                file_path = file_obj.file.path
                
                if os.path.exists(file_path):
                    ext = os.path.splitext(file_path)[1]
                    arcname = f"{file_obj.title}{ext}"
                    zip_file.write(file_path, arcname=arcname)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="favorite_files.zip"'
        
        return response


class ExportAllFilesExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        files = UserFile.objects.all().order_by('-uploaded_at')

        # Фильтры (по аналогии с AllFilesListView)
        stage = request.query_params.get('stage', '')

        if stage:
            files = files.filter(stage=stage)

        search = request.query_params.get('search', '')

        if search:
            files = files.filter(title__icontains=search)

        sort = request.query_params.get('sort', '-uploaded_at')
        files = files.order_by(sort)
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)

        # Создаём Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Файлы"

        # Заголовки
        headers = [
            "ID", "Название", "Этап", "Дата загрузки", "Версия",
            "Автор (email)", "Автор (username)", "Кто менял (email)", "Кто менял (username)", "Дата изменения"
        ]
        ws.append(headers)

        # Данные
        for f in files:
            ws.append([
                f.id,
                f.title,
                dict(UserFile.STAGE_CHOICES).get(f.stage, f.stage),
                f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                f.version,
                f.owner.email if f.owner else '',
                f.owner.username if f.owner else '',
                f.last_edited_by.email if f.last_edited_by else '',
                f.last_edited_by.username if f.last_edited_by else '',
                f.last_edited_at.strftime('%Y-%m-%d %H:%M') if f.last_edited_at else ''
            ])

        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                
            ws.column_dimensions[column].width = max_length + 2

        # Отправляем файл
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=all_files.xlsx'
        wb.save(response)

        return response


class ExportMyFilesExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        files = UserFile.objects.filter(owner=request.user).order_by('-uploaded_at')

        # Фильтры (по аналогии с UserFileListView)
        stage = request.query_params.get('stage', '')
        
        if stage:
            files = files.filter(stage=stage)
        
        search = request.query_params.get('search', '')
        
        if search:
            files = files.filter(title__icontains=search)
        
        sort = request.query_params.get('sort', '-uploaded_at')
        files = files.order_by(sort)
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)

        # Создаём Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Мои файлы"

        headers = [
            "ID", "Название", "Этап", "Дата загрузки", "Версия",
            "Автор (email)", "Автор (username)", "Кто менял (email)", "Кто менял (username)", "Дата изменения"
        ]
        ws.append(headers)

        for f in files:
            ws.append([
                f.id,
                f.title,
                dict(UserFile.STAGE_CHOICES).get(f.stage, f.stage),
                f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                f.version,
                f.owner.email if f.owner else '',
                f.owner.username if f.owner else '',
                f.last_edited_by.email if f.last_edited_by else '',
                f.last_edited_by.username if f.last_edited_by else '',
                f.last_edited_at.strftime('%Y-%m-%d %H:%M') if f.last_edited_at else ''
            ])

        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=my_files.xlsx'
        wb.save(response)

        return response


class ExportFavoriteFilesExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Получаем id избранных файлов пользователя
        favorite_ids = FavoriteFile.objects.filter(user=request.user).values_list('file_id', flat=True)
        files = UserFile.objects.filter(id__in=favorite_ids)

        # Фильтры (по аналогии с MyFavoriteFilesView)
        stage = request.query_params.get('stage', '')
        
        if stage:
            files = files.filter(stage=stage)
        
        search = request.query_params.get('search', '')
        
        if search:
            files = files.filter(title__icontains=search)
        
        sort = request.query_params.get('sort', '-uploaded_at')
        files = files.order_by(sort)
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)

        # Создаём Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Избранные файлы"

        headers = [
            "ID", "Название", "Этап", "Дата загрузки", "Версия",
            "Автор (email)", "Автор (username)", "Кто менял (email)", "Кто менял (username)", "Дата изменения"
        ]
        ws.append(headers)

        for f in files:
            ws.append([
                f.id,
                f.title,
                dict(UserFile.STAGE_CHOICES).get(f.stage, f.stage),
                f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                f.version,
                f.owner.email if f.owner else '',
                f.owner.username if f.owner else '',
                f.last_edited_by.email if f.last_edited_by else '',
                f.last_edited_by.username if f.last_edited_by else '',
                f.last_edited_at.strftime('%Y-%m-%d %H:%M') if f.last_edited_at else ''
            ])

        # Автоширина колонок
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=favorite_files.xlsx'
        wb.save(response)

        return response


# Список и создание проектов
class ProjectListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        projects = Project.objects.filter(owner=request.user).order_by('-created_at')
        serializer = ProjectSerializer(projects, many=True, context={'request': request})
        
        return Response(serializer.data)

    def post(self, request):
        name = request.data.get('name')
        
        if not name:
            return Response({'error': 'Название обязательно'}, status=400)
        
        project = Project.objects.create(owner=request.user, name=name)
        serializer = ProjectSerializer(project, context={'request': request})
        
        return Response(serializer.data, status=201)


# Получить/удалить/переименовать проект
class ProjectDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            project = Project.objects.get(pk=pk, owner=request.user)
        except Project.DoesNotExist:
            return Response({'error': 'Проект не найден'}, status=404)
        
        serializer = ProjectSerializer(project, context={'request': request})
        
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            project = Project.objects.get(pk=pk, owner=request.user)
        except Project.DoesNotExist:
            return Response({'error': 'Проект не найден'}, status=404)
        
        project.delete()
        
        return Response({'message': 'Проект удалён'}, status=204)

    def patch(self, request, pk):
        try:
            project = Project.objects.get(pk=pk, owner=request.user)
        except Project.DoesNotExist:
            return Response({'error': 'Проект не найден'}, status=404)
        
        name = request.data.get('name')
        
        if not name:
            return Response({'error': 'Название обязательно'}, status=400)
        
        project.name = name
        project.save()
        serializer = ProjectSerializer(project, context={'request': request})
        
        return Response(serializer.data)


# Добавить файл в проект
class AddFileToProjectView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, project_id):
        file_id = request.data.get('file_id')
        
        try:
            project = Project.objects.get(pk=project_id, owner=request.user)
            file = UserFile.objects.get(pk=file_id)
        except (Project.DoesNotExist, UserFile.DoesNotExist):
            return Response({'error': 'Проект или файл не найден'}, status=404)
        
        # Проверка на уникальность
        if ProjectFile.objects.filter(project=project, file=file).exists():
            return Response({'error': 'Файл уже в проекте'}, status=400)
        
        ProjectFile.objects.create(project=project, file=file)
        
        return Response({'message': 'Файл добавлен в проект'}, status=201)


# Удалить файл из проекта
class RemoveFileFromProjectView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, project_id, file_id):
        try:
            project = Project.objects.get(pk=project_id, owner=request.user)
            pf = ProjectFile.objects.get(project=project, file_id=file_id)
        except (Project.DoesNotExist, ProjectFile.DoesNotExist):
            return Response({'error': 'Проект или файл не найден'}, status=404)
        
        pf.delete()
        
        return Response({'message': 'Файл удалён из проекта'}, status=204)


class DownloadProjectFilesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            project = Project.objects.get(pk=project_id, owner=request.user)
        except Project.DoesNotExist:
            return Response({'error': 'Проект не найден'}, status=404)
        
        project_files = project.project_files.all()
        
        if not project_files:
            return Response({'error': 'Нет файлов для скачивания'}, status=404)

        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for pf in project_files:
                file_obj = pf.file
                file_path = file_obj.file.path
                
                if os.path.exists(file_path):
                    ext = os.path.splitext(file_path)[1]
                    arcname = f"{file_obj.title}{ext}"
                    zip_file.write(file_path, arcname=arcname)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="project_{project_id}_files.zip"'
        
        return response


class ExportProjectFilesExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            project = Project.objects.get(pk=project_id, owner=request.user)
        except Project.DoesNotExist:
            return Response({'error': 'Проект не найден'}, status=404)
        
        project_files = project.project_files.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Файлы проекта"

        headers = [
            "ID", "Название", "Этап", "Дата загрузки", "Версия",
            "Автор (email)", "Автор (username)", "Кто менял (email)", "Кто менял (username)", "Дата изменения"
        ]
        ws.append(headers)

        for pf in project_files:
            f = pf.file
            ws.append([
                f.id,
                f.title,
                dict(UserFile.STAGE_CHOICES).get(f.stage, f.stage),
                f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                f.version,
                f.owner.email if f.owner else '',
                f.owner.username if f.owner else '',
                f.last_edited_by.email if f.last_edited_by else '',
                f.last_edited_by.username if f.last_edited_by else '',
                f.last_edited_at.strftime('%Y-%m-%d %H:%M') if f.last_edited_at else ''
            ])

        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=project_{project_id}_files.xlsx'
        wb.save(response)

        return response
